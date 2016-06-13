# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from xlrd import open_workbook
from geo import Geojp
import csv as csv
import requests


class Main(object):

    _map = Geojp()
    _lcc = {}
    _lst = {}

    def make_lst(self):
        # --------- THIS ONE USE FOR CREATE '_lst' AND '_lcc'
        wb = load_workbook('master_data.xlsx')
        sheet1 = wb['DOM_SECTION_MST']
        for row in sheet1.iter_rows():
            if row[5].value == 1:
                if row[1].value in self._lst:
                    self._lst[row[1].value].append(row[2].value)
                else:
                    self._lst[row[1].value] = [row[2].value]

    @staticmethod
    def getprice(departure, arrival, day='20160620'):
        """Get the lowest cost of treval between 2 port"""
        url = 'http://10.10.20.132/best_price/tabicapi_api'
        dp = departure
        ar = arrival
        dp_time = day
        info = {'departure_port': dp, 'arrival_port': ar, 'departure_date': dp_time}
        r = ''
        try:
            r = requests.get(url, params=info)
        except requests.exceptions.ConnectionError:
            print 'Quỳ, ca này khó, bác sĩ bó tay !'
        else:
            while r.status_code != 200:
                r = requests.get(url, params=info)
        try:
            k = r.json()['f_price']
        except KeyError:
            return 666666
        else:
            return k

    def get_total_price(self, dep1, arr1, mid):
        """Get price for all travel path"""
        f1 = self.getprice(departure=dep1, arrival=mid)
        f2 = self.getprice(departure=mid, arrival=arr1)
        if f1 and f2:
            return f1 + f2
        else:
            return None

    @staticmethod
    def get_key(lst, get_len=3):
        rs = []
        if len(lst) <= get_len:
            for item in lst:
                rs.append(item[1])
        else:
            for idn in range(0, get_len):
                rs.append(lst[idn][1])
        return rs

    def sort_and_filter(self, start, end, listitem, lcc):
        # rs = []
        # frs = []
        rs_lcc = []
        rs_normal = []
        num = len(listitem)
        if num == 1:
            if (start in lcc and listitem[0] in lcc[start]) or (listitem[0] in lcc and end in lcc[listitem[0]]):
                return ','.join(listitem), None
            else:
                return None, ','.join(listitem)
        else:
            for item in listitem:
                if (start in lcc and item in lcc[start]) or (item in lcc and end in lcc[item]):
                    rs_lcc.append(item)
                else:
                    rs_normal.append(item)
        num_lcc = len(rs_lcc)
        if num_lcc >= 3 or num_lcc == num:
            return ','.join(rs_lcc), None
        else:
            # if num >= 3:
            #     remain = 3 - num_lcc
            # else:
            #     remain = num - num_lcc
            remain = 3 - num_lcc if num >= 3 else num - num_lcc
            rs = []
            for nm in rs_normal:
                k = self.get_total_price(dep1=start, arr1=end, mid=nm)
                if k:
                    rs.append([k, nm])
            rs.sort(key=lambda x: x[0])
            return ','.join(rs_lcc), ','.join(self.get_key(rs, remain))

    def get_way(self, depart, desti, dictaw):
        """Function for calculate continous airport"""
        out = []  # mảng chứ giá trị kết quả
        con = []  # mảng tạm chứa giá trị trung chuyển
        if depart in dictaw:
            con = dictaw[depart]
        des_area, des_id = self._map.get_area(desti)
        dep_area, dep_id = self._map.get_area(depart)
        if dep_id < des_id:
            for item in con:
                if (dep_id <= self._map.get_area_id(item) <= des_id) and (desti in dictaw[item]):
                    out.append(item)
                else:
                    pass
        elif dep_id > des_id:
            for item in con:
                if (dep_id >= self._map.get_area_id(item) >= des_id) and (desti in dictaw[item]):
                    out.append(item)
                else:
                    pass
        else:
            for item in con:
                if ((des_id - 1) <= self._map.get_area_id(item) <= (des_id + 1)) and (desti in dictaw[item]):
                    out.append(item)
                else:
                    pass
        return out\


    def get_price_lcc(self, dep, arr):
        # TODO
        return None

    def sort_list_lcc(self, str_lst_lcc):
        lst_lcc = list(str_lst_lcc.split(','))
        k = len(lst_lcc)
        if k == 1:
            return str_lst_lcc
        else:
            tmp = []
            for item in lst_lcc:
                tmp.append([self.get_price_lcc(item), item])
            tmp.sort(key=lambda x:x[0])
            return ','.join(self.get_key(tmp, k))

    def sort_lcc(self):
        """Sort cac chuyen lcc xem chuyen nao re hon"""
        print '--- running sort lcc ---'
        wb = open_workbook('final_result_2.xlsx')
        sheet = wb.sheet_by_name('DOM_SECTION_MST')

        csv_file = open('result_trial.csv', 'wb')
        outputwritter = csv.writer(csv_file, delimiter='\t')
        outputwritter.writerow(['DEPARTURE', 'ARRIVAL', 'LCC_PORT_ORDINAL'])

        start_row = 1
        end_row = 150
        none_break = True
        for index in range(sheet.nrows):
            print 'runing row --- ', index
            if index > end_row:
                break
            elif start_row <= index <= end_row:
                row = sheet.row(index)
                if row[9].value:
                    outputwritter.writerow([row[1].value, row[2].value, self.sort_list_lcc(row[9].value)])
            else:
                print '------- start '





    def run(self):

        # dep = 'AOJ'
        # des = 'RIS'
        print '--------------------------------------------------'
        # print dep, _map.get_area_id(dep)
        # print des, _map.get_area_id(des)
        # print _map.get_area('HND')
        # print _map.get_area('NRT')
        # for key, value in _lst.iteritems():
        #     print key, value
        # k = get_way(depart=dep, desti=des, dictaw=_lst)
        # print k
        # print ','.join(_lst[dep])
        # print ','.join(_lst['CTS'])
        # print ','.join(get_way(depart=dep, desti=des, dictaw=_lst))
        # print 'AKJ -> HSG : ', get_way('AKJ', 'ITM', lst)
        # nwb = Workbook()
        # sheet = nwb.create_sheet(title='RESULT')
        # sheet = wb['DOM_SECTION_MST']

        # --------- MAKE ROUTES BY THIS ONE -------------
        # -------------- THIS ONE DONE ------------

        # for row in sheet1.iter_rows():
        #     if row[4].value:
        #         k = get_way(depart=row[1].value, desti=row[2].value, dictaw=_lst)
        #         if k:
        #             index = 'I' + str(row[0].value + 1)
        #             # sheet1[index].value = k.__str__()
        #             sheet1[index].value = ','.join(k)
        #
        # sheet1['I1'].value = 'RESULT'
        # wb.save('result.xlsx')

        # -------------- THIS ONE FOR FILTER ---------
        # --------------- Run for 100 row in excel ---------------
        #
        #
        # wb = load_workbook('result.xlsx')
        # sheet = wb['DOM_SECTION_MST']
        # start_row = 1
        # end_row = 2450
        # sheet['J1'] = 'FINAL_RESULT'
        # for index, row in enumerate(sheet.iter_rows()):
        #     if start_row <= index <= end_row:
        #         print 'row', index, 'run ...'
        #         ind = 'J' + str(index + 1)
        #         if row[8].value:
        #             tmp_lst = list(row[8].value.split(','))
        #             if len(tmp_lst) != 1:
        #                 sheet[ind].value = ','.join(sort_and_filter(start=row[1].value, end=row[2].value, listitem=tmp_lst))
        #             else:
        #                 sheet[ind].value = ','.join(tmp_lst)
        #         print 'finish', index
        # wb.save('final_result.xlsx')
        # # --------------------------------------------------------

        wb = load_workbook('final_result.xlsx')
        sheet = wb['DOM_SECTION_MST']
        start_row = 1
        end_row = 2415  # 2415 for full

        # ----- WILL REMOVE TO TOP WHEN FINISH FINAL ------
        for row in sheet.iter_rows():
            if row[6].value == 1:
                if row[1].value in self._lcc:
                    self._lcc[row[1].value].append(row[2].value)
                else:
                    self._lcc[row[1].value] = [row[2].value]

        for index, row in enumerate(sheet.iter_rows()):
            if start_row <= index <= end_row:
                print 'row', index, 'run ...'
                inj = 'J' + str(index + 1)
                ink = 'K' + str(index + 1)
                if row[8].value:
                    tmp_lst = list(row[8].value.split(','))
                    sheet[inj].value, sheet[ink] = self.sort_and_filter(start=row[1].value,
                                                                        end=row[2].value,
                                                                        listitem=tmp_lst,
                                                                        lcc=self._lcc)
                print 'finish', index
        # sheet['K1'] = 'NORMAL_RESULTS'
        # sheet['J1'] = 'LCC_RESULTS'
        for items in self._lcc.iteritems():
            print(items)
        wb.save('final_result_2.xlsx')
        print 'Successful ! THIS IS SPARTA'
